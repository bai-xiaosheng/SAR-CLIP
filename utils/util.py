import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import openpyxl
import scipy.io as scio
from sklearn.metrics import confusion_matrix
from PIL.Image import Image
from torchvision.transforms import Compose, Resize, CenterCrop, ToTensor, Normalize,ToPILImage
import torch.nn.functional as F
try:
    from torchvision.transforms import InterpolationMode
    BICUBIC = InterpolationMode.BICUBIC
except ImportError:
    BICUBIC = Image.BICUBIC
#存储结果
def create_file(file_path, file_name='exp'):
    folder = os.path.exists(file_path)
    if not folder:
        os.makedirs(file_path)
    num = 1
    file_list = os.listdir(file_path)
    for i in file_list:
        if os.path.exists(file_path + './%s' % (file_name + '_' + str(num))):
            num += 1
    os.makedirs(file_path + './%s' % (file_name + '_' + str(num)))
    path = os.path.join(file_path, (file_name + '_%d' % num))
    # exp = file_name + '_' + str(num)
    print('result will save:', path)
    return path
#创建csv文件
def creat_csv(path,model='Train'):
    if model == 'Train':
        datafile = pd.DataFrame(columns=['epoch','train_loss','train_acc'])
        path = os.path.join(path,'train_result.csv')
        datafile.to_csv(path, index=False)
    elif model == 'Test':
        datafile = pd.DataFrame(columns=['epoch','test_acc'])
        path = os.path.join(path,'test_result.csv')
        datafile.to_csv(path, index=False)
    return path
#保存结果csv
def save_result(data,path):
    data=pd.DataFrame([data])
    data.to_csv(path,mode='a',header=False,index=False)
#
def creat_excel(path,sheet_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(path)
#
def write_excel(path,value,append=True,columns=0):
    if not os.path.exists(path):
        creat_excel(path,'sheet1')
    # value = value.values
    data = openpyxl.load_workbook(path)
    #
    sheetnames = data.sheetnames
    sheet = data[sheetnames[0]]
    sheet = data.active
    if append:
        startrows = sheet.max_row
    else:
        startrows = 0
    for i in range(len(value)):
        for j in range(len(value[i])):
            sheet.cell(row=startrows+i+1,column=columns+j+1,value=str(value[i][j]))
    data.save(path)
    # print('实验结果保存完成！')

#显示最优结果
def show_result(path):
    data = pd.read_csv(path)
    test_acc = np.array(data[['test_acc']])
    # test_epoch = np.array(data[['epoch']])
    max_acc = np.max(test_acc)
    i = np.argmax(test_acc)
    print("epoch:{} max_acc={:.2f}".format(i,max_acc))

#画损失曲线
def draw_loss(path):
    data = pd.read_csv(path)
    train_loss = np.array(data[['train_loss']])
    train_acc = np.array(data[['train_acc']])
    epoch = np.array(data[['epoch']])
    #准确率曲线
    plt.figure('train_acc')
    plt.plot(epoch,train_acc)  #,label='train_acc'
    plt.title('train_acc')
    plt.xlabel('epoch')
    plt.ylabel('train_acc')
    # plt.legend()
    #损失函数曲线
    plt.figure('train_loss')
    plt.plot(epoch,train_loss)
    plt.xlabel('epoch')
    plt.ylabel('train_loss')
    plt.title('train_loss')
    # plt.legend()
    plt.show()
#画注意力机制
def plot_attention(a,b,c):

    plt.subplot(2, 2, 1)
    plt.imshow(a)
    plt.subplot(2, 2, 2)
    plt.imshow(b, cmap='gray')
    plt.subplot(2, 2, 3)
    plt.imshow(a - b)
    plt.subplot(2, 2, 4)
    plt.imshow(c)
    plt.show()

#对图像进行裁剪、转换为RGB图像
def _convert_image_to_rgb(image):
    return image.convert("RGB")
def data_preprocess(x,n_px=(224,224)):
    tranfer = Compose([
        # CenterCrop((60,60)),
        Resize(n_px, interpolation=BICUBIC),  # , interpolation=BICUBIC
        # CenterCrop(n_px),
        _convert_image_to_rgb,
        ToTensor(),
        # Normalize((0.48145466, 0.4578275, 0.40821073), (0.26862954, 0.26130258, 0.27577711)),
        # Normalize((0.48145466, 0.48145466, 0.48145466), (0.26862954, 0.26862954, 0.26862954)),
    ])
    # x.show()
    x = tranfer(x)
    # plt.imshow(x.permute(1,2,0))
    # plt.show()
    # x.show()
    return x

def draw_confusion_matrix(label_true, label_pred, save_path,title="Confusion Matrix", dpi=100):
    # label_true = torch.Tensor(label_true,device='cpu')
    # label_pred = torch.Tensor(label_pred,device='cpu')
    labels_true = []
    labels_pred = []
    for i in range(len(label_true)):
        labels_true = labels_true + label_true[i].cpu().numpy().tolist()
        labels_pred = labels_pred + label_pred[i].cpu().numpy().tolist()

    labels_true = np.array(labels_true).flatten()

    labels_pred = np.array(labels_pred).flatten()

    c = confusion_matrix(y_true=labels_true, y_pred=labels_pred, normalize='true')
    print('C is\n',c)
    num = 1
    for i in os.listdir(save_path):
        if os.path.exists(save_path+'./%s'%('test_'+str(num)+'.mat')):
            num +=1
    epo = 'test_'+str(num)+'.mat'
    dataFile = os.path.join(save_path,epo)
    scio.savemat(dataFile,{'C':c})
    plt.show()
    return dataFile

def calculate_metric(logic,label,correct=0):
    pred = logic.max(1,keepdim=True)[1].squeeze()
    loss = F.cross_entropy(logic,label)
    correct += pred.eq(label).sum().item()
    return pred,loss,correct

def cal_data(data):
    mean = np.mean(data)
    std = np.std(data)
    max = np.max(data)
    min = np.min(data)
    return mean,std,max,min